from flask import Flask,request,render_template,Response
import openpyxl
from selenium import webdriver
import Barcode_gen
import time
import json
import os
base_path=r"E:\Hackerearth\NPCI\DataB"
database_path=base_path+r"\Person_db.xlsx"
toll_booth_path=base_path+r"\toll_booth_db.xlsx"
hash_base=base_path+r"\user_hash.json"
read_hash=r"E:\Hackerearth\NPCI\code_scan\hash.txt"
scanned_file=r"E:\Hackerearth\NPCI\code_scan\scanned.jpg"
bar=r"E:\Hackerearth\NPCI\code_scan\abc.png"
app=Flask(__name__)
info={}

def validate(name,pwd):
    wb = openpyxl.load_workbook(database_path)
    sh = wb['User']
    check=None
    index=None
    for i in range(1,50):
        if name ==sh.cell(row=i,column=1).value:
            if pwd==sh.cell(row=i,column=3).value:
                info['name']=name
                info['vehicle_id']=sh.cell(row=i,column=2).value
                wb.close()
                return True
    wb.close()
    return False

def store():
    print('store')
    wb = openpyxl.load_workbook(toll_booth_path)
    print('wb opened')
    print(wb.sheetnames)
    sh=wb['Toll_booth']
    lc=1
    print(sh.cell(row=lc, column=1).value)
    while(True):
        if(sh.cell(row=lc,column=1).value==None):
            break
        else:
            lc+=1
    print(lc)
    flg='NEW_ENTRY'
    for i in range(lc,1,-1):
        if(sh.cell(row=i,column=1).value==info['vehicle_id'] and sh.cell(row=i, column=5).value.lower() != 'n'):
            cur=time.time()
            tdiff=cur-float(sh.cell(row=i, column=3).value)
            print(tdiff)
            if (tdiff/86400>=1):
                sh.cell(row=i, column=5).value = 'N'
                flg='NEW_ENTRY'
                break
            else:
                flg='RETURN_TRIP'
                sh.cell(row=i, column=5).value = 'N'
                sh.cell(row=i, column=4).value = 'Done'
                wb.save(toll_booth_path)
                break


    if(flg=='RETURN_TRIP'):
        print('ReturnEntry')
        return(True)

    if(flg=='NEW_ENTRY'):
        print('NewEntry')
        sh.cell(row=lc,column=1).value = info['vehicle_id']
        print('v_id')
        sh.cell(row=lc, column=2).value = info['toll_name']
        print('t_nm')
        sh.cell(row=lc, column=3).value = info['intime']
        print('intim')
        sh.cell(row=lc, column=5).value = 'Y'
        print('act')
        wb.save(toll_booth_path)
        return(True)

def hash_match():
    print('hash_match')
    global scan_flg
    print('hm',scan_flg)
    while(not os.path.exists(scanned_file)):
        pass
    with open(read_hash,'r') as f:
        hsh=f.read()
        print('hsh', hsh)
        if hsh==info['hash']:
            scan_flg=True
            os.remove(scanned_file)
            return(True)
    scan_flg = True
def barc():
    time_ = time.time()
    info['intime'] = time_
    ar = [info['name'], info['toll_name'], info['vehicle_id'], str(info['intime'])]
    print(ar)
    hash = Barcode_gen.encrypt(ar)
    print(hash)
    Barcode_gen.generate_barcode(hash)
    dat=None
    with open(hash_base,'r') as f:
        dat=json.load(f)
        dat[info['vehicle_id']]=hash
        info['hash']=hash
    with open(hash_base,'w') as f:
        json.dump(dat,f,indent=2)


@app.route('/user/')
def hello_():
    return render_template('login.html')

# @app.route('/scanner/')
# def scan():
#     return

@app.route('/form_login',methods=['POST','GET'])
def login():

    name=request.form['un']
    pwd=request.form['pwd']
    print(name,pwd)
    st=validate(name,pwd)
    if st==True:
        # driver2.refresh()
        return render_template('toll_name.html')
    else:
        return "Invalid"

@app.route('/toll_name',methods=['POST','GET'])
def toll_name():

    toll_name=request.form['toll_name']
    info['toll_name']=toll_name
    barc()
    return render_template('barcode_disp.html')

@app.route('/barcode_disp', methods=['POST', 'GET'])
def barcode_disp():
    print('barcode_disp')
    if(hash_match()):
        sts=store()
        if(sts==True):
            return render_template('status.html')

    else:
        return 'INVALID'


@app.route('/scanner/',methods=['POST','GET'])
def loop():
    return render_template('scanner.html')

@app.route('/video_feed')
def video_feed():
    global scan_flg
    if os.path.exists(scanned_file):
        scan_flg=False
    """Video streaming route. Put this in the src attribute of an img tag."""
    return Response(Barcode_gen.read_feed(scan_flg),
                    mimetype='multipart/x-mixed-replace; boundary=frame')


if __name__=='__main__':
    global driver
    global driver2
    driver=webdriver.Chrome()
    driver.set_window_position(0, 0)
    driver.set_window_size(480, 640)
    driver.get('http://127.0.0.1:5000/user/')
    global scan_flg
    scan_flg = True
    driver2=webdriver.Chrome()
    driver2.set_window_position(500,0)
    driver2.set_window_size(480, 640)
    driver2.get('http://127.0.0.1:5000/scanner/')
    app.run()
