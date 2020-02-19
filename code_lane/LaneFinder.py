from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import selenium.webdriver.support.ui as ui
import time
import openpyxl
import constants

#################
lane_inp_page=constants.lane_inp_page
database_path=constants.database_path
schema=constants.schema
schema_begin=constants.schema_begin
schema_end=constants.schema_end
lane_result_page=constants.lane_result_page
#################

def create_script(info):
    cnt=info['lane_count']
    stats=info['stats']
    bgin=None
    end=None
    with open(schema_begin,'r') as f:
        bgin=f.readlines()
    print(bgin)
    with open(schema_end,'r') as f:
        end=f.readlines()
    with open(lane_result_page,'w') as f:
        f.writelines(bgin)
        for i in range(len(stats)):
            f.write(schema.format(lane_num=stats[i][0],pay_type=stats[i][1],active=stats[i][2]))
        f.writelines(end)


def get_details(name):
    wb = openpyxl.load_workbook(database_path)
    sh=wb['data']
    index=None
    info={'name':name}
    for i in range(3,50):
        t_nam=sh.cell(row=i, column=1).value
        # print(sh.cell(row=i,column=1).value)
        if(t_nam==name):
            index=i
            break
    info['code'] = sh.cell(row=index,column=2).value
    info['lane_count'] = sh.cell(row=index, column=3).value
    info['stats'] = []
    col=3
    for i in range(1,int(info['lane_count'])+1):
        col+=1
        pay=sh.cell(row=index, column=col).value
        col += 1
        active=sh.cell(row=index, column=col).value
        info['stats'].append([i,pay,active])
    return(info)

if __name__=="__main__":
    global driver
    driver = webdriver.Chrome()
    driver.set_window_position(0,0)
    driver.set_window_size(480,640)
    driver.get(lane_inp_page)
    inp=driver.find_element_by_id('inp')
    time.sleep(3)
    # inp.send_keys("Aluva")
    sub=driver.find_element_by_id("submit")
    toll_name=''
    while True:
        if (sub.get_attribute("value")!='GO'):
            toll_name=inp.get_attribute("value")
            # print(inp.get_attribute("value"))
            break
    time.sleep(1)
    # driver.close()
    info=get_details(toll_name)

    # info=get_details('Hebbalu')
    print(info)
    create_script(info)
    driver.get(lane_result_page)